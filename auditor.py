import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill

def ejecutar_auditoria(ruta_archivo_entrada, ruta_archivo_salida):
    df = pd.read_excel(ruta_archivo_entrada)
    
    def auditoria_completa(fila):
        relato = str(fila.get('relato', '')).upper()
        caratula_cruda = str(fila.get('calificaciones', '')).upper()
        caratula_analisis = str(fila.get('analisis_caratula', '')).strip().upper()
        modalidad_analisis = str(fila.get('analisis_modalidad', '')).strip().upper()
        armas_analisis = str(fila.get('analisis_armas', '')).strip().upper()
        lugar_analisis = str(fila.get('analisis_lugar', '')).strip().upper()
        
        # 1. Validación de Coordenadas
        lat = fila.get('analisis_coordenadas', None)
        if pd.isna(lat) or str(lat).strip() == "":
            lat = fila.get('latitud', None)
            
        if pd.notna(lat) and str(lat).strip() != "":
            try:
                lat_str = str(lat).strip()
                if ',' in lat_str and '-' in lat_str[1:]:
                    partes = lat_str.split(',')
                    lat_str = partes[0].strip()
                lat_str = lat_str.replace(',', '.')
                lat_num = float(lat_str)
                if not (-42 <= lat_num <= -33):
                    return "ERROR: Coordenadas fuera de rango geográfico provincial"
            except Exception:
                return "ERROR: Formato de coordenadas inválido"
        else:
            return "ERROR: Falta registrar coordenadas"

        # 2. Excepciones para carátulas descartables
        caratulas_descartables = ['LESIONES LEVES', 'LESIONES CULPOSAS', 'AVERIGUACION DE PARADERO', 'DAÑOS', 'INCENDIO', 'ESTRAGO', 'VIOLACION DE DOMICILIO']
        es_descartable = any(item in caratula_cruda for item in caratulas_descartables)
        analisis_vacio = pd.isna(fila.get('analisis_caratula')) or caratula_analisis == "" or caratula_analisis == "NAN"
        
        if 'VIOLACION DE DOMICILIO' in caratula_cruda or 'DOMICILIO' in caratula_cruda:
            if 'BICICLETA' in relato and 'HURTO' not in caratula_analisis and 'ROBO' not in caratula_analisis:
                return "ERROR: El relato indica sustracción (debe reclasificarse como Hurto o Robo)"

        if es_descartable and analisis_vacio:
            return "CORRECTO"
            
        if analisis_vacio:
            return "ERROR: Falta completar la carátula analizada"

        # 3. Control de Lugar (Vía pública vs Domicilio/Comercio)
        menciona_casa_o_finca = any(term in relato for term in ["EN SU DOMICILIO", "INTERIOR DE SU DOMICILIO", "CASA", "DEPARTAMENTO", "FRENTE A SU DOMICILIO", "PATIO", "LOCAL", "COMERCIO"])
        if menciona_casa_o_finca and "VIA PUBLICA" in lugar_analisis:
            return "ALERTA: El relato indica lugar cerrado/domicilio/comercio pero se cargó Vía Pública"

        # 3.1. REGLA ESTRICTA DE DIFERENCIACIÓN: Comercio vs Finca
        menciona_comercio = any(term in relato for term in ["LOCAL", "COMERCIO", "NEGOCIO", "KIOSCO", "SUPERMERCADO", "FARMACIA", "LOCAL COMERCIAL"])
        menciona_vivienda = any(term in relato for term in ["CASA", "DEPARTAMENTO", "VIVIENDA", "DOMICILIO PARTICULAR"])
        
        if menciona_comercio and ("FINCA" in caratula_analisis or "FINCA" in modalidad_analisis):
            return "ALERTA: El relato menciona un comercio pero se tipificó como Finca"
            
        if menciona_vivienda and ("COMERCIO" in modalidad_analisis or "LOCAL" in modalidad_analisis):
            return "ALERTA: El relato menciona una vivienda pero se tipificó como Comercio"

        # 4. Reglas para Hurtos
        if "HURTO" in caratula_analisis:
            if "ROBA RUEDAS" in modalidad_analisis and not ("RUEDA" in relato or "NEUMATICO" in relato or "AUXILIO" in relato):
                return "ALERTA: Modalidad Roba Ruedas pero el relato no menciona neumáticos"
            if "BICICLETA" in modalidad_analisis and "BICICLETA" not in relato:
                return "ALERTA: Modalidad Bicicleta pero el relato no menciona rodado"
            return "CORRECTO"

        if "ESTUPEFACIENTES" in caratula_analisis:
            return "CORRECTO"

        # 5. Validación estricta de Armas (Afinada para evitar falsos positivos)
        palabras_relato = re.findall(r'\b\w+\b', relato)
        menciona_arma_real = False
        
        terminos_armas = ["PISTOLA", "REVOLVER", "ESCOPETA", "CUCHILLO", "NAVAJA", "AMENAZA CON ARMA", "ARMA DE FUEGO", "ARMA BLANCA"]
        
        if any(term in relato for term in terminos_armas) or ("ARMA" in palabras_relato and "SIN" not in relato and "NO" not in relato):
            if not any(neg in relato for neg in ["SIN ARMA", "NO PORTABA", "CARECE DE ARMA", "SIN EL EMPLEO DE ARMA", "SIN EXHIBICION"]):
                menciona_arma_real = True

        tipos_armas_validos = ['FUEGO', 'BLANCA', 'IMPROPIA']
        if menciona_arma_real and not any(tipo in armas_analisis for tipo in tipos_armas_validos):
            return "ERROR: El relato menciona un arma real pero no se tipificó correctamente"

        # 6. Robos y modalidades
        if "ROBO" in caratula_analisis:
            if "MOTOCHORRO" in modalidad_analisis and not ("MOTO" in relato or "MOTOCICLETA" in relato):
                return "ALERTA: Modalidad Motochorro pero el relato no menciona moto"
            if "ESCRUCHE" in modalidad_analisis or "FINCA" in modalidad_analisis:
                if not any(term in relato for term in ["FORZAR", "ROTURA", "CANDADO", "VENTANA", "PUERTA", "AUSENTES", "SIN MORADORES"]):
                    return "ALERTA: Modalidad Finca/Escruche pero el relato no menciona forzamiento o ausencia"

        # 7. Sustracción Vehicular por Levantamiento
        if "SUSTRACCIÓN AUTOMOTOR" in caratula_analisis or "SUSTRACCION MOTOVEHICULO" in caratula_analisis:
            if "LEVANTAMIENTO" in modalidad_analisis:
                return "CORRECTO"

        return "CORRECTO"

    df['ESTADO_AUDITORIA_VISUAL'] = df.apply(auditoria_completa, axis=1)
    
    with pd.ExcelWriter(ruta_archivo_salida, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Control')
        
    wb = openpyxl.load_workbook(ruta_archivo_salida)
    ws = wb.active
    
    verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    rojo = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    amarillo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    col_idx = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'ESTADO_AUDITORIA_VISUAL':
            col_idx = col
            break
            
    if col_idx:
        for row in range(2, ws.max_row + 1):
            celda = ws.cell(row=row, column=col_idx)
            val = str(celda.value)
            if val == 'CORRECTO':
                celda.fill = verde
            elif 'ERROR' in val:
                celda.fill = rojo
            elif 'ALERTA' in val:
                celda.fill = amarillo
                
    wb.save(ruta_archivo_salida)
    print("Auditoría completada exitosamente.")

if __name__ == "__main__":
    ejecutar_auditoria("denuncias_julio.xlsx", "resultado_auditoria.xlsx")
